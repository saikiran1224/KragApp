from pathlib import Path 
from pypdf import PdfReader
from openpyxl import load_workbook
from docx import Document

# Function to determine the file extension
def determine_file_extension(file_path):
    ext = Path(file_path).suffix.lower() 
    file_name = Path(file_path).name
    return ext, file_name # returns .pdf, .xlsx, etc... and file_name

# Reader function for PDF - Private 
def _load_pdf(file_path): 

    reader = PdfReader(file_path) # loading the pdf 
    text = "" 
    for page in reader.pages: 
        text += page.extract_text() 

    return text # returning the pdf content in form of raw string 

# Reader function for XLSX - Private 
def _load_xlsx(file_path):
    wb = load_workbook(file_path)
    text = ""
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            text += " ".join([str(cell) for cell in row if cell is not None]) + "\n"
    
    return text # returning the xlsx content in form of raw string

# Reader function for TXT - Private
def _load_txt(file_path):
    with open(file_path, "r", encoding="utf-8") as f:
        text = f.read()
    return text # returns the text content in form of raw string 


# Reader function for DOCX - Private
def _load_docx(file_path):

    doc = Document(file_path)
    text = "\n".join([para.text for para in doc.paragraphs])

    return text # returns the document content in form of raw string


# Defining all the available loaders in one Dispatch table
LOADERS = {
    ".pdf" : _load_pdf, 
    ".xlsx" : _load_xlsx,
    ".txt" : _load_txt,
    ".docx" : _load_docx
}

# Main Router Function - Entrypoint
def load_document(file_path):

    # Step 1 - Determining the file type 
    extension, file_name = determine_file_extension(file_path)

    # Step 2 - Identifying the loader 
    loader_identified = LOADERS.get(extension)

    if not loader_identified: # If no suitable loader found
        raise ValueError(f"Unsupported file type: {extension}. Only .pdf, .xlsx, .docx and .txt files are Accepted.")

    # Step 3 - Fetching the raw text content using the help of loader by passing the file_path as the parameter
    text = loader_identified(file_path)

    # Step 4 - Building the JSON format to return 
    return {
        "text" : text, 
        "file_name" : file_name,
        "file_type" : extension.lstrip("."),
        "metadata" : {}
    }

    